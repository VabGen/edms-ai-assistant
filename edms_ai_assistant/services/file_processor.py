"""
EDMS AI Assistant - Enhanced File Processor Service.

Поддерживает: PDF, DOCX, DOC (legacy Word 97-2003), TXT, XLSX, XLS.
PDF со сканами распознаётся через OCR (PyMuPDF + pytesseract) с кэшированием на диск.
Внедряет маркеры страниц для интеллектуального поиска по документу.
"""

from __future__ import annotations

import asyncio
import io
import logging
import os
import shutil
import subprocess
from concurrent.futures import ThreadPoolExecutor
from functools import lru_cache
from pathlib import Path
from typing import Any

from langchain_community.document_loaders import TextLoader

logger = logging.getLogger(__name__)

_MIN_AVG_CHARS_PER_PAGE = 20
_CACHE_SUFFIX = ".ocr_cache.txt"


# ── Lazy Tesseract Discovery ─────────────────────────────────


@lru_cache(maxsize=1)
def _get_tesseract_cmd() -> str:
    """Ищет бинарник Tesseract один раз и кэширует результат."""
    env_path = os.getenv("TESSERACT_CMD")
    if env_path and Path(env_path).exists():
        return env_path

    path = shutil.which("tesseract")
    if path:
        return path

    win_paths = [
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
    ]
    for p in win_paths:
        if Path(p).exists():
            return p

    raise RuntimeError(
        "Tesseract OCR binary not found. "
        "Install tesseract-ocr or set TESSERACT_CMD env var."
    )


@lru_cache(maxsize=1)
def _get_tessdata_prefix() -> str | None:
    """Определяет путь к tessdata для Tesseract OCR (лениво)."""
    env_prefix = os.getenv("TESSDATA_PREFIX")
    if env_prefix and Path(env_prefix).is_dir():
        return env_prefix

    linux_paths = [
        "/usr/share/tesseract-ocr/5/tessdata",
        "/usr/share/tesseract-ocr/4.00/tessdata",
        "/usr/share/tessdata",
        "/usr/local/share/tessdata",
    ]
    for p in linux_paths:
        if Path(p).is_dir() and list(Path(p).glob("*.traineddata")):
            return p

    win_path = r"C:\Program Files\Tesseract-OCR\tessdata"
    if Path(win_path).is_dir():
        return win_path

    return None


@lru_cache(maxsize=1)
def _get_available_ocr_langs() -> list[str]:
    """Проверяет доступные языки для Tesseract OCR."""
    prefix = _get_tessdata_prefix()
    if not prefix:
        return []

    path = Path(prefix)
    langs = []
    for lang in ["rus", "eng"]:
        if (path / f"{lang}.traineddata").exists():
            langs.append(lang)
    return langs


# ── Utility: OCR Disk Cache ─────────────────────────────────────────────


def _get_cached_ocr(file_path: str) -> str | None:
    """Возвращает кэшированный OCR текст, если исходный файл не изменился."""
    cache_path = file_path + _CACHE_SUFFIX
    src_path = Path(file_path)

    if not Path(cache_path).exists():
        return None

    try:
        src_stat = src_path.stat()
        cache_stat = Path(cache_path).stat()

        if src_stat.st_mtime > cache_stat.st_mtime or src_stat.st_size == 0:
            return None

        with open(cache_path, encoding="utf-8") as f:
            text = f.read()
            if text.strip():
                logger.info("Using cached OCR result for %s", file_path)
                return text
    except Exception as e:
        logger.warning("Failed to read OCR cache: %s", e)
    return None


def _save_ocr_cache(file_path: str, text: str) -> None:
    """Сохраняет распознанный текст в кэш на диск."""
    cache_path = file_path + _CACHE_SUFFIX
    try:
        with open(cache_path, "w", encoding="utf-8") as f:
            f.write(text)
    except Exception as e:
        logger.warning("Failed to save OCR cache: %s", e)


# ── Extractors ──────────────────────────────────────────────────────────


def _extract_doc_via_fitz(file_path: str) -> str:
    """Extract text from .doc or .docx using PyMuPDF (fitz)."""
    import fitz  # type: ignore[import]

    with fitz.open(file_path) as doc:
        pages_text = [
            page.get_text("text").strip()
            for page in doc
            if page.get_text("text").strip()
        ]
    return "\n\n".join(pages_text)


def _extract_doc_via_mammoth(file_path: str) -> str:
    """Extract text from .doc/.docx using mammoth library as fallback."""
    import mammoth  # type: ignore[import]

    with open(file_path, "rb") as f:
        result = mammoth.extract_raw_text(f)
    return result.value.strip()


def _extract_docx_via_docx2txt(file_path: str) -> str:
    """Extract text from .docx using docx2txt (ZIP-based format only)."""
    import docx2txt  # type: ignore[import]

    return docx2txt.process(file_path) or ""


# ── Utility: DOC to DOCX conversion ─────────────────────────────────────


def _convert_doc_to_docx(doc_path: str) -> str:
    """Convert legacy .doc (Word 97-2003) to modern .docx using LibreOffice."""
    doc_path_obj = Path(doc_path)
    out_dir = doc_path_obj.parent

    soffice_cmd = shutil.which("soffice")
    if not soffice_cmd:
        win_paths = [
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        ]
        for p in win_paths:
            if Path(p).exists():
                soffice_cmd = p
                break

    if not soffice_cmd:
        raise RuntimeError(
            "LibreOffice not found. Install from https://www.libreoffice.org/ "
            "and ensure 'soffice' is in PATH or specify full path."
        )

    try:
        subprocess.run(
            [
                soffice_cmd,
                "--headless",
                "--convert-to",
                "docx:MS Word 2007 XML",
                "--outdir",
                str(out_dir),
                str(doc_path_obj),
            ],
            capture_output=True,
            text=True,
            timeout=60,
            check=True,
        )
        converted_path = out_dir / f"{doc_path_obj.stem}.docx"
        if converted_path.exists():
            logger.info(
                "DOC converted to DOCX via LibreOffice: %s -> %s",
                doc_path,
                converted_path,
            )
            return str(converted_path)
        raise RuntimeError(f"Converted file not found: {converted_path}")

    except FileNotFoundError as e:
        raise RuntimeError(f"LibreOffice not found or failed to start: {e}")
    except subprocess.TimeoutExpired:
        raise RuntimeError(f"LibreOffice conversion timed out for: {doc_path}")
    except subprocess.CalledProcessError as e:
        raise RuntimeError(f"Conversion failed: {e.stderr or e}")


# ── Main Service ────────────────────────────────────────────────


class FileProcessorService:
    """Service for text extraction from multiple file formats.

    Extraction strategies by extension:
        .pdf    -> fitz Text Layer -> OCR fallback (fitz + pytesseract) + Disk Cache
        .docx   -> docx2txt -> mammoth -> fitz (fallback chain)
        .doc    -> fitz (PyMuPDF) -> mammoth -> LibreOffice (fallback chain)
        .txt    -> TextLoader (LangChain)
        .xlsx   -> openpyxl
        .xls    -> xlrd
    """

    SUPPORTED_EXTENSIONS = {
        ".pdf",
        ".docx",
        ".doc",
        ".txt",
        ".xlsx",
        ".xls",
    }

    _default_instance: FileProcessorService | None = None

    def __init__(self, max_workers: int = 4):
        self._executor = ThreadPoolExecutor(max_workers=max_workers)

    @classmethod
    def _get_default_instance(cls) -> FileProcessorService:
        """Ленивая инициализация синглтона для обратной совместимости с @classmethod."""
        if cls._default_instance is None:
            cls._default_instance = cls()
        return cls._default_instance

    @classmethod
    async def extract_text_async(cls, file_path: str, executor: ThreadPoolExecutor | None = None) -> str:
        """Async text extraction delegating CPU work to thread pool."""
        loop = asyncio.get_running_loop()
        exec_pool = executor or cls._get_default_instance()._executor
        return await loop.run_in_executor(exec_pool, cls.extract_text, file_path)

    @classmethod
    def extract_text(cls, file_path: str) -> str:
        """Synchronous text extraction with format-specific strategy."""
        path = Path(file_path)

        if not path.exists():
            error_msg = f"Файл не найден по пути: {file_path}"
            logger.error(error_msg, extra={"file_path": file_path})
            return f"Ошибка: {error_msg}"

        ext = path.suffix.lower()

        if ext not in cls.SUPPORTED_EXTENSIONS:
            warning_msg = f"Формат файла {ext} пока не поддерживается для анализа."
            logger.warning(
                "Unsupported file format: %s", ext, extra={"file_path": file_path}
            )
            return warning_msg

        try:
            if ext in (".xlsx", ".xls"):
                return cls._extract_from_excel(file_path, ext)
            if ext == ".doc":
                return cls._extract_doc(file_path)
            if ext == ".docx":
                return cls._extract_docx(file_path)
            if ext == ".pdf":
                return cls._extract_pdf(file_path)
            if ext == ".txt":
                return cls._extract_txt(file_path)

            return f"Формат {ext} не поддерживается."

        except Exception as e:
            error_msg = f"Произошла техническая ошибка при чтении файла {ext}: {e!s}"
            logger.error(
                "File parsing error: %s",
                e,
                exc_info=True,
                extra={"file_path": file_path},
            )
            return error_msg

    # ── Format-specific extractors ────────────────────────────────────────────

    @classmethod
    def _extract_pdf(cls, file_path: str) -> str:
        """Unified PDF pipeline: Cache -> Text Layer (fitz) -> OCR."""
        cached = _get_cached_ocr(file_path)
        if cached:
            return cached

        import fitz  # type: ignore[import]

        with fitz.open(file_path) as doc:
            pages_text = []
            page_count = len(doc)

            for i, page in enumerate(doc):
                text = page.get_text("text").strip()
                if text:
                    pages_text.append(f"--- Страница {i + 1} ---\n{text}")

            full_text = "\n\n".join(pages_text)

        avg_chars = len(full_text) / max(page_count, 1)

        if full_text and avg_chars >= _MIN_AVG_CHARS_PER_PAGE:
            logger.info(
                "PDF extracted via text layer (fitz)",
                extra={"chars": len(full_text), "pages": page_count},
            )
            _save_ocr_cache(file_path, full_text)
            return full_text

        logger.info(
            "PDF text layer too thin (%d chars / %d pages ≈ %.0f chars/page), switching to OCR",
            len(full_text),
            page_count,
            avg_chars,
        )

        # Fallback на OCR
        try:
            ocr_text = cls._extract_pdf_via_ocr(file_path)
            if ocr_text and ocr_text.strip():
                _save_ocr_cache(file_path, ocr_text)
                return ocr_text
        except ImportError as ie:
            logger.warning("OCR dependencies missing: %s", ie)
        except RuntimeError as re_err:
            logger.warning("OCR runtime error: %s", re_err)
        except Exception as e:
            logger.error("OCR extraction failed: %s", e, exc_info=True)

        return (
            "Не удалось извлечь текст из PDF.\n"
            "Документ содержит только изображения без текстового слоя.\n\n"
            "Для распознавания сканов установите Tesseract OCR или переменную TESSERACT_CMD."
        )

    @classmethod
    def _extract_pdf_via_ocr(cls, file_path: str) -> str:
        """Extract text from scanned (image-only) PDF using PyMuPDF + Tesseract OCR."""
        import fitz  # type: ignore[import]
        import pytesseract  # type: ignore[import]
        from PIL import Image  # type: ignore[import]

        # ── Настройка pytesseract (ленивая) ──────────────────────────────
        tesseract_cmd = _get_tesseract_cmd()
        pytesseract.pytesseract.tesseract_cmd = tesseract_cmd

        if not os.getenv("TESSDATA_PREFIX"):
            prefix = _get_tessdata_prefix()
            if prefix:
                os.environ["TESSDATA_PREFIX"] = prefix

        # Определяем доступные языки
        available_langs = _get_available_ocr_langs()
        if "rus" in available_langs and "eng" in available_langs:
            ocr_lang = "rus+eng"
        elif "eng" in available_langs:
            logger.warning("Russian OCR language not available, using eng only")
            ocr_lang = "eng"
        elif "rus" in available_langs:
            logger.warning("English OCR language not available, using rus only")
            ocr_lang = "rus"
        else:
            raise RuntimeError(
                "No OCR languages available. Install tesseract-ocr-rus and tesseract-ocr-eng."
            )

        logger.info(
            "Starting PDF OCR: lang=%s, tesseract=%s",
            ocr_lang,
            tesseract_cmd,
        )

        # ── Извлечение текста постранично ────────────────────────────────
        pages_text: list[str] = []

        with fitz.open(file_path) as doc:
            total_pages = len(doc)
            for page_num in range(total_pages):
                page = doc[page_num]

                zoom = 300 / 72
                mat = fitz.Matrix(zoom, zoom)
                pix = page.get_pixmap(matrix=mat, alpha=False)

                img = None
                try:
                    img_bytes = pix.tobytes("png")
                    img = Image.open(io.BytesIO(img_bytes))

                    try:
                        text = pytesseract.image_to_string(img, lang=ocr_lang)
                    except pytesseract.TesseractError as te:
                        logger.warning(
                            "Tesseract failed on page %d with lang=%s: %s — trying 'eng' only",
                            page_num + 1,
                            ocr_lang,
                            te,
                        )
                        if ocr_lang != "eng":
                            try:
                                text = pytesseract.image_to_string(img, lang="eng")
                            except pytesseract.TesseractError as te2:
                                logger.error(
                                    "Tesseract eng-only also failed on page %d: %s",
                                    page_num + 1,
                                    te2,
                                )
                                text = ""
                        else:
                            text = ""

                    if text and text.strip():
                        pages_text.append(
                            f"--- Страница {page_num + 1} ---\n{text.strip()}"
                        )
                finally:
                    if img:
                        img.close()
                    pix = None

        logger.info(
            "PDF OCR completed: %d pages, %d with text",
            total_pages,
            len(pages_text),
        )
        return "\n\n".join(pages_text)

    @classmethod
    def _extract_doc(cls, file_path: str) -> str:
        """Extract text from legacy .doc (Word 97-2003) files."""
        try:
            text = _extract_doc_via_fitz(file_path)
            if text and len(text.strip()) > 10:
                logger.info(
                    "DOC extracted via fitz",
                    extra={"file_path": file_path, "chars": len(text)},
                )
                return text
        except Exception as fitz_err:
            logger.warning(
                "fitz failed for .doc '%s': %s — trying mammoth", file_path, fitz_err
            )

        try:
            text = _extract_doc_via_mammoth(file_path)
            if text and len(text.strip()) > 10:
                logger.info(
                    "DOC extracted via mammoth",
                    extra={"file_path": file_path, "chars": len(text)},
                )
                return text
        except ImportError:
            logger.warning("mammoth not installed — install with: uv add mammoth")
        except Exception as mammoth_err:
            logger.warning("mammoth failed for .doc '%s': %s", file_path, mammoth_err)

        try:
            logger.info("Attempting LibreOffice conversion for .doc: %s", file_path)
            docx_path = _convert_doc_to_docx(file_path)
            text = cls._extract_docx(docx_path)
            try:
                Path(docx_path).unlink()
            except Exception as cleanup_err:
                logger.warning(
                    "Failed to remove temp file %s: %s", docx_path, cleanup_err
                )

            if text and len(text.strip()) > 10:
                logger.info(
                    "DOC extracted via LibreOffice conversion",
                    extra={"original": file_path, "chars": len(text)},
                )
                return text

        except RuntimeError as conv_err:
            logger.warning(
                "LibreOffice conversion failed for '%s': %s", file_path, conv_err
            )
        except Exception as e:
            logger.error(
                "Unexpected error during DOC extraction for '%s': %s",
                file_path,
                e,
                exc_info=True,
            )

        return (
            "Не удалось извлечь текст из файла .doc.\n"
            "Возможные причины:\n"
            "• Файл повреждён или в неподдерживаемом бинарном формате.\n"
            "• LibreOffice не установлен (требуется для конвертации).\n"
            "Рекомендация: пересохраните документ в формате .docx."
        )

    @classmethod
    def _extract_docx(cls, file_path: str) -> str:
        """Extract text from .docx (Office Open XML ZIP) files."""
        try:
            text = _extract_docx_via_docx2txt(file_path)
            if text and len(text.strip()) > 10:
                logger.info(
                    "DOCX extracted via docx2txt",
                    extra={"file_path": file_path, "chars": len(text)},
                )
                return text
        except KeyError as ke:
            logger.warning("docx2txt failed (invalid ZIP) for '%s': %s", file_path, ke)
        except Exception as e:
            logger.warning("docx2txt failed for '%s': %s", file_path, e)

        try:
            text = _extract_doc_via_mammoth(file_path)
            if text and len(text.strip()) > 10:
                logger.info(
                    "DOCX extracted via mammoth",
                    extra={"file_path": file_path, "chars": len(text)},
                )
                return text
        except Exception as e:
            logger.warning("mammoth failed for '%s': %s — trying fitz", file_path, e)

        try:
            text = _extract_doc_via_fitz(file_path)
            if text and len(text.strip()) > 10:
                logger.info(
                    "DOCX extracted via fitz",
                    extra={"file_path": file_path, "chars": len(text)},
                )
                return text
        except Exception as e:
            logger.error("fitz also failed for '%s': %s", file_path, e)

        return "Не удалось извлечь текст из файла .docx. Файл может быть повреждён."

    @classmethod
    def _extract_txt(cls, file_path: str) -> str:
        """Extract text from plain .txt files."""
        loader = TextLoader(file_path, encoding="utf-8")
        docs = loader.load()
        if not docs:
            return "Текстовый файл пуст."
        full_text = "\n\n".join(doc.page_content for doc in docs).strip()
        logger.info(
            "TXT extracted successfully",
            extra={"file_path": file_path, "chars": len(full_text)},
        )
        return full_text

    @classmethod
    def _extract_from_excel(cls, file_path: str, ext: str) -> str:
        """Extract text from Excel files preserving table structure."""
        try:
            if ext == ".xlsx":
                import openpyxl

                wb = openpyxl.load_workbook(file_path, data_only=True)
            else:
                import xlrd

                wb = xlrd.open_workbook(file_path)

            extracted_text = []

            if ext == ".xlsx":
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    extracted_text.append(
                        f"\n{'=' * 50}\nЛИСТ: {sheet_name}\n{'=' * 50}\n"
                    )
                    for row in sheet.iter_rows(values_only=True):
                        if any(cell is not None for cell in row):
                            row_text = " | ".join(
                                str(cell) if cell is not None else "" for cell in row
                            )
                            extracted_text.append(row_text)
            else:
                for sheet_idx in range(wb.nsheets):
                    sheet = wb.sheet_by_index(sheet_idx)
                    extracted_text.append(
                        f"\n{'=' * 50}\nЛИСТ: {sheet.name}\n{'=' * 50}\n"
                    )
                    for row_idx in range(sheet.nrows):
                        row = sheet.row_values(row_idx)
                        if any(cell for cell in row):
                            row_text = " | ".join(
                                str(cell) if cell else "" for cell in row
                            )
                            extracted_text.append(row_text)

            result = "\n".join(extracted_text).strip()
            logger.info(
                "Excel extracted successfully",
                extra={"file_path": file_path, "extension": ext, "chars": len(result)},
            )
            return result

        except ImportError as e:
            logger.error("Missing dependency for Excel: %s", e)
            return "Ошибка: Для обработки Excel файлов требуется установить 'openpyxl' или 'xlrd'."
        except Exception as e:
            logger.error("Excel extraction error: %s", e, exc_info=True)
            return f"Ошибка при чтении Excel файла: {e!s}"

    # ── Structured data / utility methods ────────────────────────────────

    @classmethod
    async def extract_structured_data(cls, file_path: str) -> dict[str, Any]:
        """Extract structured data including text, metadata, and tables."""
        path = Path(file_path)
        ext = path.suffix.lower()
        text = await cls.extract_text_async(file_path)
        stats = {
            "chars": len(text),
            "words": len(text.split()),
            "lines": text.count("\n"),
            "digits": len([c for c in text if c.isdigit()]),
        }
        file_stat = path.stat()
        metadata = {
            "filename": path.name,
            "extension": ext,
            "size_bytes": file_stat.st_size,
            "size_mb": round(file_stat.st_size / (1024 * 1024), 2),
        }
        result: dict[str, Any] = {
            "text": text,
            "metadata": metadata,
            "stats": stats,
            "tables": None,
        }
        if ext in (".xlsx", ".xls"):
            result["tables"] = await cls._extract_excel_tables(file_path, ext)
        return result

    @classmethod
    async def _extract_excel_tables(
            cls, file_path: str, ext: str
    ) -> list[dict[str, Any]]:
        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(
            cls._get_default_instance()._executor, cls._extract_tables_sync, file_path, ext
        )

    @classmethod
    def _extract_tables_sync(cls, file_path: str, ext: str) -> list[dict[str, Any]]:
        try:
            if ext == ".xlsx":
                import openpyxl

                wb = openpyxl.load_workbook(file_path, data_only=True)
            else:
                import xlrd

                wb = xlrd.open_workbook(file_path)
            tables = []
            if ext == ".xlsx":
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    data = [
                        list(row)
                        for row in sheet.iter_rows(values_only=True)
                        if any(c is not None for c in row)
                    ]
                    if data:
                        tables.append(
                            {
                                "sheet_name": sheet_name,
                                "headers": data[0],
                                "data": data[1:],
                                "rows_count": len(data) - 1,
                            }
                        )
            else:
                for i in range(wb.nsheets):
                    sheet = wb.sheet_by_index(i)
                    data = [
                        sheet.row_values(r)
                        for r in range(sheet.nrows)
                        if any(sheet.row_values(r))
                    ]
                    if data:
                        tables.append(
                            {
                                "sheet_name": sheet.name,
                                "headers": data[0],
                                "data": data[1:],
                                "rows_count": len(data) - 1,
                            }
                        )
            return tables
        except Exception as e:
            logger.error("Table extraction error: %s", e, exc_info=True)
            return []

    @classmethod
    def validate_file_path(cls, file_path: str) -> str | None:
        if not file_path or not file_path.strip():
            return "Путь к файлу не может быть пустым"
        path = Path(file_path)
        if not path.exists():
            return f"Файл не найден: {file_path}"
        if not path.is_file():
            return f"Указанный путь не является файлом: {file_path}"
        ext = path.suffix.lower()
        if ext not in cls.SUPPORTED_EXTENSIONS:
            supported = ", ".join(sorted(cls.SUPPORTED_EXTENSIONS))
            return f"Неподдерживаемый формат {ext}. Поддерживаются: {supported}"
        return None

    @classmethod
    def get_file_info(cls, file_path: str) -> dict[str, Any]:
        path = Path(file_path)
        if not path.exists():
            return {"exists": False, "error": f"Файл не найден: {file_path}"}
        try:
            stat = path.stat()
            return {
                "exists": True,
                "name": path.name,
                "extension": path.suffix.lower(),
                "size_bytes": stat.st_size,
                "size_mb": round(stat.st_size / (1024 * 1024), 2),
                "is_supported": path.suffix.lower() in cls.SUPPORTED_EXTENSIONS,
                "absolute_path": str(path.absolute()),
            }
        except Exception as e:
            return {"exists": True, "error": f"Не удалось получить информацию: {e!s}"}

    @classmethod
    def get_ocr_diagnostic(cls) -> dict[str, Any]:
        """Возвращает диагностику Tesseract OCR для API /health и отладки."""
        result: dict[str, Any] = {
            "binary_found": False,
            "binary_path": None,
            "tessdata_found": False,
            "tessdata_path": None,
            "languages": [],
            "version": None,
            "errors": [],
        }

        try:
            cmd = _get_tesseract_cmd()
            result["binary_found"] = True
            result["binary_path"] = cmd

            proc = subprocess.run(
                [cmd, "--version"],
                capture_output=True,
                text=True,
                timeout=10,
            )
            version_line = (proc.stdout or proc.stderr or "").strip().split("\n")[0]
            result["version"] = version_line
        except RuntimeError as e:
            result["errors"].append(str(e))
        except Exception as e:
            result["errors"].append(f"Failed to run tesseract --version: {e}")

        prefix = _get_tessdata_prefix()
        result["tessdata_found"] = prefix is not None
        result["tessdata_path"] = prefix

        if not prefix:
            result["errors"].append(
                "tessdata directory not found. Install languages or set TESSDATA_PREFIX env var."
            )
        else:
            result["languages"] = _get_available_ocr_langs()

        return result
